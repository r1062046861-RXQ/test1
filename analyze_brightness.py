import os
import numpy as np
from PIL import Image

ASSETS_DIR = os.path.join(os.path.dirname(__file__), 'game', 'public', 'assets')

# All background images with their usage info
background_images = [
    {
        'file': 'background_main_menu.png',
        'usage': 'IntroView / StartMenu / EventView(x2) / ChestView / CardCodexView / MapView(后备)'
    },
    {
        'file': 'background_map_act1.png',
        'usage': 'MapView (Act 1)'
    },
    {
        'file': 'background_map_act2.png',
        'usage': 'MapView (Act 2)'
    },
    {
        'file': 'background_map.png',
        'usage': '仅在 manifest 中声明，代码中未使用'
    },
    {
        'file': 'background_combat_act1.png',
        'usage': 'CombatView (Act 1 / fallback)'
    },
    {
        'file': 'background_combat_act2.png',
        'usage': 'CombatView (Act 2)'
    },
    {
        'file': 'background_combat_act3.png',
        'usage': 'CombatView (Act 3)'
    },
    {
        'file': 'background_rest.png',
        'usage': 'RestView'
    },
    {
        'file': 'background_shop.png',
        'usage': 'ShopView'
    },
    {
        'file': 'bg_synthesis_1.png',
        'usage': 'SynthesisBench (随机)'
    },
    {
        'file': 'bg_synthesis_2.png',
        'usage': 'SynthesisBench (随机)'
    },
]

def compute_brightness_stats(image_path):
    """Compute luminance statistics for an image.
    Luminance formula: Y = 0.299*R + 0.587*G + 0.114*B
    Returns dict with mean, median, min, max, std, and brightness category.
    """
    img = Image.open(image_path).convert('RGB')
    arr = np.array(img, dtype=np.float64)
    
    # Perceived luminance (ITU-R BT.601)
    luminance = 0.299 * arr[:,:,0] + 0.587 * arr[:,:,1] + 0.114 * arr[:,:,2]
    
    return {
        'width': img.width,
        'height': img.height,
        'pixels': img.width * img.height,
        'mean': float(np.mean(luminance)),
        'median': float(np.median(luminance)),
        'min': float(np.min(luminance)),
        'max': float(np.max(luminance)),
        'std': float(np.std(luminance)),
    }

def brightness_level(mean_val):
    if mean_val < 30:
        return '极暗 (Very Dark)'
    elif mean_val < 60:
        return '暗 (Dark)'
    elif mean_val < 100:
        return '偏暗 (Dim)'
    elif mean_val < 140:
        return '中等 (Medium)'
    elif mean_val < 180:
        return '偏亮 (Bright)'
    elif mean_val < 220:
        return '亮 (Very Bright)'
    else:
        return '极亮 (Ultra Bright)'

def main():
    results = []
    
    for bg in background_images:
        filepath = os.path.join(ASSETS_DIR, bg['file'])
        if not os.path.exists(filepath):
            print(f'[WARN] File not found: {filepath}')
            continue
        
        stats = compute_brightness_stats(filepath)
        results.append({
            **bg,
            **stats,
            'level': brightness_level(stats['mean'])
        })
        print(f'Done: {bg["file"]} - Mean Luminance: {stats["mean"]:.1f}')
    
    # Print summary table
    print('\n' + '=' * 120)
    print(f'{"文件名":<35} {"分辨率":<16} {"均值":>7} {"中位数":>7} {"最小值":>7} {"最大值":>7} {"标准差":>7} {"亮度评级"}')
    print('=' * 120)
    for r in results:
        res = f'{r["width"]}x{r["height"]}'
        print(f'{r["file"]:<35} {res:<16} {r["mean"]:>7.2f} {r["median"]:>7.2f} {r["min"]:>7.2f} {r["max"]:>7.2f} {r["std"]:>7.2f}  {r["level"]}')
    
    # Generate Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = '背景图亮度数据'
    
    # Styles
    header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Consolas', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )
    
    # Color scales for brightness
    very_dark_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    dark_fill = PatternFill(start_color='2D2D44', end_color='2D2D44', fill_type='solid')
    dim_fill = PatternFill(start_color='4A4A6A', end_color='4A4A6A', fill_type='solid')
    medium_fill = PatternFill(start_color='7A7A9A', end_color='7A7A9A', fill_type='solid')
    bright_fill = PatternFill(start_color='B0B0C8', end_color='B0B0C8', fill_type='solid')
    very_bright_fill = PatternFill(start_color='E0E0F0', end_color='E0E0F0', fill_type='solid')
    ultra_bright_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    dark_text = Font(name='Consolas', size=11, color='FFFFFF', bold=True)
    light_text = Font(name='Consolas', size=11, color='333333', bold=True)
    
    fill_map = {
        '极暗 (Very Dark)': (very_dark_fill, dark_text),
        '暗 (Dark)': (dark_fill, dark_text),
        '偏暗 (Dim)': (dim_fill, dark_text),
        '中等 (Medium)': (medium_fill, light_text),
        '偏亮 (Bright)': (bright_fill, light_text),
        '亮 (Very Bright)': (very_bright_fill, light_text),
        '极亮 (Ultra Bright)': (ultra_bright_fill, light_text),
    }
    
    # Column headers
    headers = [
        '序号', '背景图文件名', '分辨率', '像素总数',
        '平均亮度\n(Mean)', '中位数亮度\n(Median)',
        '最小亮度\n(Min)', '最大亮度\n(Max)',
        '标准差\n(Std)', '亮度评级',
        '使用界面'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for i, r in enumerate(results):
        row = i + 2
        res_str = f'{r["width"]}x{r["height"]}'
        row_data = [
            i + 1,
            r['file'],
            res_str,
            r['pixels'],
            round(r['mean'], 2),
            round(r['median'], 2),
            round(r['min'], 2),
            round(r['max'], 2),
            round(r['std'], 2),
            r['level'],
            r['usage'],
        ]
        
        level_fill, level_font = fill_map.get(r['level'], (None, data_font))
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col in (1, 3, 4, 5, 6, 7, 8, 9):  # numeric columns center
                cell.alignment = data_alignment
                cell.font = data_font
            elif col == 10:  # brightness level with color
                cell.alignment = data_alignment
                cell.font = level_font
                cell.fill = level_fill
            else:
                cell.alignment = left_alignment
                cell.font = data_font
    
    # Column widths
    col_widths = [6, 30, 14, 14, 14, 14, 14, 14, 14, 24, 52]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Row height
    ws.row_dimensions[1].height = 36
    for row in range(2, len(results) + 2):
        ws.row_dimensions[row].height = 22
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = f'A1:K{len(results) + 1}'
    
    # ---- Summary sheet ----
    ws2 = wb.create_sheet('汇总统计')
    
    summary_data = [
        ['指标', '值'],
        ['总背景图数量', len(results)],
        ['平均亮度均值', round(np.mean([r['mean'] for r in results]), 2)],
        ['亮度均值范围', f'{min(r["mean"] for r in results):.2f} ~ {max(r["mean"] for r in results):.2f}'],
        ['平均标准差', round(np.mean([r['std'] for r in results]), 2)],
        ['最暗背景图', min(results, key=lambda x: x['mean'])['file']],
        ['最亮背景图', max(results, key=lambda x: x['mean'])['file']],
        ['', ''],
        ['亮度分布:', ''],
    ]
    
    # Count per level
    from collections import Counter
    level_counts = Counter(r['level'] for r in results)
    for level, count in level_counts.most_common():
        summary_data.append([level, f'{count} 张'])
    
    for row, (label, val) in enumerate(summary_data, 1):
        c1 = ws2.cell(row=row, column=1, value=label)
        c2 = ws2.cell(row=row, column=2, value=val)
        if row == 1:
            c1.font = header_font
            c1.fill = header_fill
            c1.alignment = header_alignment
            c2.font = header_font
            c2.fill = header_fill
            c2.alignment = header_alignment
        else:
            c1.font = Font(name='Microsoft YaHei', size=11, bold=True)
            c1.alignment = Alignment(horizontal='right', vertical='center')
            c2.font = Font(name='Consolas', size=11)
            c2.alignment = Alignment(horizontal='left', vertical='center')
        c1.border = thin_border
        c2.border = thin_border
    
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 40
    ws2.freeze_panes = 'A2'
    
    # Save
    output_path = os.path.join(os.path.dirname(__file__), 'background_brightness_report.xlsx')
    wb.save(output_path)
    print(f'\nExcel report saved to: {output_path}')
    return results

if __name__ == '__main__':
    main()
