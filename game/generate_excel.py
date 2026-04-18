import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def create_card_master_excel():
    # Define data based on design.txt and current codebase
    # Using the IDs and properties from design.txt (which aligns with cards.ts mostly)
    # Adding special "cards" like Constitutions and Enemies as requested.
    
    # Headers
    headers = [
        "序号", "名称", "占位图名称", "功能描述", "出现位置", "类型", "稀有度", "费用", "备注"
    ]
    
    # Data List
    data = []
    
    # --- 1. Player Cards (From design.txt) ---
    # We will map the design doc IDs (86+) to the game logic IDs if possible, or just use sequential.
    # The user asked for "ID (int)". Let's start from 1.
    
    # Helper to add row
    def add_card(idx, name, img, desc, loc, ctype, rarity, cost, note=""):
        data.append([idx, name, img, desc, loc, ctype, rarity, cost, note])

    current_id = 1
    
    # 1.1 Balanced (Pinghe)
    add_card(current_id, "陈皮理气", "86", "抽2张牌，丢弃1张牌。", "PlayerDeck", "Skill", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "丹参活血", "87", "造成7点伤害。给予目标2层血瘀。", "PlayerDeck", "Attack", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "山楂消食", "88", "造成3点伤害。如果击杀敌人，获得5点格挡。", "PlayerDeck", "Attack", "Common", 0, "平和质核心"); current_id += 1
    add_card(current_id, "薏苡除湿", "89", "获得4点格挡，并移除自身的1个负面状态。", "PlayerDeck", "Skill", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "生姜发散", "90", "本回合，你的下一张攻击卡额外造成3点伤害。消耗。", "PlayerDeck", "Skill", "Common", 0, "平和质核心"); current_id += 1
    add_card(current_id, "金银花露", "91", "对所有敌人造成3点伤害，并清除每个敌人的1层正面状态。", "PlayerDeck", "Skill", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "桂枝通络", "92", "获得3点格挡，本回合你的攻击卡无视敌人3点格挡。", "PlayerDeck", "Skill", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "白芍柔肝", "93", "使一个敌人虚弱2回合(造成的伤害减少25%)，抽1张牌。", "PlayerDeck", "Skill", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "川芎行气", "94", "造成5点伤害，抽1张牌。", "PlayerDeck", "Attack", "Common", 1, "平和质核心"); current_id += 1
    add_card(current_id, "小柴胡汤", "95", "恢复5点生命，抽2张牌，获得4点格挡。", "PlayerDeck", "Skill", "Uncommon", 2, "平和质核心"); current_id += 1

    # 1.2 Yin Deficiency
    add_card(current_id, "麦冬滋阴", "96", "获得2层滋阴。抽1张牌。", "PlayerDeck", "Skill", "Uncommon", 1, "阴虚质核心"); current_id += 1
    add_card(current_id, "生地凉血", "97", "获得4层滋阴。消耗。", "PlayerDeck", "Skill", "Uncommon", 2, "阴虚质核心"); current_id += 1
    add_card(current_id, "知母清热", "98", "获得1层滋阴。本回合你的攻击卡额外施加1层虚热。", "PlayerDeck", "Skill", "Uncommon", 1, "阴虚质核心"); current_id += 1
    add_card(current_id, "玄参泻火", "99", "消耗所有滋阴。每消耗1层，对一个随机敌人造成3点伤害。", "PlayerDeck", "Skill", "Uncommon", 1, "阴虚质核心"); current_id += 1
    add_card(current_id, "玉竹生津", "100", "(能力)在你的回合开始时，如果你有滋阴，获得1点真气。", "PlayerDeck", "Power", "Rare", 2, "阴虚质核心"); current_id += 1
    add_card(current_id, "百合安神", "101", "移除自身的1个负面状态。如果你有至少3层滋阴，改为移除所有负面状态。", "PlayerDeck", "Skill", "Uncommon", 1, "阴虚质核心"); current_id += 1
    add_card(current_id, "鳖甲软坚", "102", "获得5点格挡。你每有1层滋阴，额外获得1点格挡。", "PlayerDeck", "Skill", "Uncommon", 2, "阴虚质核心"); current_id += 1
    add_card(current_id, "石斛益胃", "103", "获得1层滋阴。本场战斗中，你的滋阴层数上限+2。消耗。", "PlayerDeck", "Skill", "Rare", 0, "阴虚质核心"); current_id += 1
    add_card(current_id, "山萸肉固涩", "104", "恢复3点生命。额外恢复等同于你滋阴层数的生命值。", "PlayerDeck", "Skill", "Uncommon", 1, "阴虚质核心"); current_id += 1
    add_card(current_id, "清骨散", "105", "造成8点伤害。消耗3层滋阴，使此卡伤害翻倍。", "PlayerDeck", "Attack", "Rare", 2, "阴虚质核心"); current_id += 1

    # 1.3 Qi Deficiency
    add_card(current_id, "黄芪固表", "106", "获得10点护盾。", "PlayerDeck", "Skill", "Uncommon", 2, "气虚质核心"); current_id += 1
    add_card(current_id, "白术健脾", "107", "获得6点格挡。如果本回合你未受到攻击伤害，获得2点力量。", "PlayerDeck", "Skill", "Uncommon", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "甘草和中", "108", "恢复4点生命，抽1张牌。", "PlayerDeck", "Skill", "Uncommon", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "山药平补", "109", "(能力)在你的回合结束时，恢复2点生命。", "PlayerDeck", "Power", "Rare", 2, "气虚质核心"); current_id += 1
    add_card(current_id, "党参补气", "110", "获得4点格挡。下回合，你的第一张技能卡效果增加2点。", "PlayerDeck", "Skill", "Uncommon", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "大枣养血", "111", "恢复2点生命，获得2点护盾。消耗。", "PlayerDeck", "Skill", "Common", 0, "气虚质核心"); current_id += 1
    add_card(current_id, "防风祛风", "112", "获得7点格挡。赋予1个敌人易伤(下次受到伤害增加50%)。", "PlayerDeck", "Skill", "Uncommon", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "茯苓渗湿", "113", "获得5点格挡。抽1张牌。消耗1层自身湿邪(如有)。", "PlayerDeck", "Skill", "Common", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "升麻升提", "114", "本回合，你每获得1次格挡或护盾，获得1点临时力量。", "PlayerDeck", "Skill", "Uncommon", 1, "气虚质核心"); current_id += 1
    add_card(current_id, "补中益气汤", "115", "(能力)每当你获得格挡时，获得等量的护盾。消耗。", "PlayerDeck", "Power", "Rare", 3, "气虚质核心"); current_id += 1

    # 2. General Herbs
    add_card(current_id, "麻黄发汗", "116", "造成8点伤害。如果敌人有寒邪，额外造成4点伤害。", "PlayerDeck", "Attack", "Common", 1, "通用药材"); current_id += 1
    add_card(current_id, "大黄攻下", "117", "造成12点伤害，给予敌人2层泄下(回合结束时失去生命)。", "PlayerDeck", "Attack", "Uncommon", 2, "通用药材"); current_id += 1
    add_card(current_id, "附子回阳", "118", "获得1点力量和1点敏捷，获得5点格挡。消耗。", "PlayerDeck", "Skill", "Rare", 3, "通用药材"); current_id += 1
    add_card(current_id, "当归补血", "119", "恢复5点生命。如果生命已满，改为获得5点护盾。", "PlayerDeck", "Skill", "Rare", 1, "通用药材"); current_id += 1
    add_card(current_id, "朱砂安神", "120", "(能力)你打出的攻击卡有25%几率使敌人晕眩(下回合无法行动)。", "PlayerDeck", "Power", "Rare", 1, "通用药材"); current_id += 1
    add_card(current_id, "黄连燥湿", "121", "造成6点伤害，并清除目标1层正面状态。", "PlayerDeck", "Attack", "Common", 1, "通用药材"); current_id += 1
    add_card(current_id, "桂枝汤", "122", "恢复8点生命，获得5点格挡。", "PlayerDeck", "Skill", "Uncommon", 2, "通用药材"); current_id += 1
    add_card(current_id, "板蓝根清热", "123", "给予所有敌人2层热邪，抽1张牌。", "PlayerDeck", "Skill", "Common", 1, "通用药材"); current_id += 1
    add_card(current_id, "三七化瘀", "124", "造成5点伤害。如果目标有血瘀，则额外造成5点伤害。", "PlayerDeck", "Attack", "Uncommon", 1, "通用药材"); current_id += 1
    add_card(current_id, "艾叶温经", "125", "获得3点格挡。本回合你每打出一张卡牌，获得1点格挡。", "PlayerDeck", "Skill", "Uncommon", 1, "通用药材"); current_id += 1
    add_card(current_id, "酸枣仁安眠", "126", "使一个敌人获得1层困倦(下回合跳过行动)。", "PlayerDeck", "Skill", "Rare", 1, "通用药材"); current_id += 1
    add_card(current_id, "泽泻利水", "127", "移除自身所有湿邪，每移除1层，获得2点格挡。", "PlayerDeck", "Skill", "Uncommon", 1, "通用药材"); current_id += 1
    add_card(current_id, "枳实行气", "128", "抽1张牌。如果你本回合已打出一张攻击卡，额外抽1张牌。", "PlayerDeck", "Skill", "Common", 0, "通用药材"); current_id += 1
    add_card(current_id, "连翘解毒", "129", "对所有敌人造成4点伤害，并清除1层热邪。", "PlayerDeck", "Skill", "Common", 1, "通用药材"); current_id += 1
    add_card(current_id, "肉桂引火", "130", "获得3点力量。本回合结束时，失去3点力量。", "PlayerDeck", "Skill", "Uncommon", 2, "通用药材"); current_id += 1

    # 3. Prescriptions (Consumables)
    add_card(current_id, "麻黄汤", "131", "本回合，你的攻击卡无视敌人所有格挡。", "Shop", "Skill", "Common", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "四君子汤", "132", "恢复15点生命，并获得5点格挡。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "逍遥散", "133", "清除所有负面状态。抽3张牌。", "Shop", "Skill", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "安宫牛黄丸", "134", "使你在本场战斗中免疫下一次死亡，并恢复20%生命。", "Shop", "Power", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "六味地黄丸", "135", "获得5层滋阴和5点护盾。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "银翘散", "136", "对所有敌人造成8点伤害，并施加1层热邪。", "Shop", "Attack", "Common", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "血府逐瘀汤", "137", "清除所有敌人所有正面状态。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "藿香正气散", "138", "清除自身所有负面状态，并恢复10点生命。", "Shop", "Skill", "Common", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "龙胆泻肝汤", "139", "对一个敌人造成20点真实伤害(无视格挡)。", "Shop", "Attack", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "玉屏风散", "140", "获得20点护盾。", "Shop", "Skill", "Common", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "归脾汤", "141", "恢复20点生命，并抽2张牌。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "乌梅丸", "142", "偷取一个敌人所有正面状态。", "Shop", "Skill", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "真武汤", "143", "本场战斗中，你的格挡效果翻倍。", "Shop", "Power", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "大承气汤", "144", "对一个敌人造成其当前生命值30%的伤害。", "Shop", "Attack", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "黄连解毒汤", "145", "对所有敌人造成10点伤害，并清除其所有正面状态。", "Shop", "Attack", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "半夏白术天麻汤", "146", "使所有敌人眩晕2回合。", "Shop", "Skill", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "金匮肾气丸", "147", "获得3点力量和3点敏捷，恢复10点生命。", "Shop", "Power", "Rare", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "清营汤", "148", "清除场上所有热邪状态，每清除1层，对所有敌人造成2点伤害。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "保和丸", "149", "抽牌直到手牌达到5张。", "Shop", "Skill", "Common", 0, "方剂(一次性)"); current_id += 1
    add_card(current_id, "川芎茶调散", "150", "本回合，你的卡牌消耗减少1点(最低为0)。", "Shop", "Skill", "Uncommon", 0, "方剂(一次性)"); current_id += 1

    # 4. Acupuncture (Events/Special)
    add_card(current_id, "针刺：足三里", "151", "(能力)每当你打出攻击卡时，恢复1点生命。", "Event", "Power", "Rare", 2, "针灸"); current_id += 1
    add_card(current_id, "艾灸：关元", "152", "获得1点真气上限。恢复3点生命。", "Event", "Skill", "Rare", 2, "针灸"); current_id += 1
    add_card(current_id, "推拿：捏脊", "153", "移除自身的2个负面状态。抽1张牌。", "Event", "Skill", "Common", 1, "针灸"); current_id += 1
    add_card(current_id, "拔罐：走罐", "154", "对所有敌人造成5点伤害。", "Event", "Attack", "Common", 1, "针灸"); current_id += 1
    add_card(current_id, "耳穴：神门", "155", "获得5点格挡。本回合，敌人对你造成的第一次伤害减少3点。", "Event", "Skill", "Common", 0, "针灸"); current_id += 1
    add_card(current_id, "刮痧：大椎", "156", "移除自身所有热邪和寒邪。", "Event", "Skill", "Uncommon", 1, "针灸"); current_id += 1
    add_card(current_id, "温针灸：三阴交", "157", "(能力)在你的回合结束时，如果你有格挡，则格挡值不会消失。", "Event", "Power", "Rare", 1, "针灸"); current_id += 1
    add_card(current_id, "点穴：合谷", "158", "使一个敌人虚弱2回合。", "Event", "Skill", "Uncommon", 1, "针灸"); current_id += 1
    add_card(current_id, "雷火灸：命门", "159", "获得3点力量和5点护盾。", "Event", "Skill", "Rare", 3, "针灸"); current_id += 1
    add_card(current_id, "砭石：阿是穴", "160", "复制自身的一个正面状态。消耗。", "Event", "Skill", "Rare", 0, "针灸"); current_id += 1

    # 5. Status Cards (Debuffs/Buffs)
    add_card(current_id, "肝火旺", "161", "(敌方能力)每回合开始，自身攻击力+2。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "脾虚湿困", "162", "(敌方能力)玩家的卡牌消耗增加1点。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "风寒束表", "163", "(敌方技能)对玩家施加1层寒邪和1层虚弱。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "热入营血", "164", "(敌方能力)回合结束时，对玩家造成等同于其热邪层数的伤害。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "肾不纳气", "165", "(敌方技能)偷取玩家1点真气上限。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "气滞血瘀", "166", "(敌方能力)玩家每打出一张牌，受到1点伤害。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "心肾不交", "167", "(敌方技能)使玩家下回合无法获得护盾和格挡。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "痰蒙心窍", "168", "(敌方技能)使玩家晕眩1回合。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "阳明腑实", "169", "(敌方能力)每回合结束时，清除玩家所有的护盾。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1
    add_card(current_id, "冲任不固", "170", "(敌方技能)使玩家失去所有正面状态。", "EnemyDeck", "Status", "Common", 0, "敌方状态"); current_id += 1

    # --- 6. Special Cards: Constitutions (Initial Avatars) ---
    add_card(current_id, "平和质", "pinghezhi", "阴阳气血调和，体态适中。无特殊效果。", "All", "Special", "Common", 0, "初始体质"); current_id += 1
    add_card(current_id, "阴虚质", "yinxuzhi", "阴虚火旺：回合开始获得+1能量，受伤害+2。", "All", "Special", "Common", 0, "初始体质"); current_id += 1
    add_card(current_id, "气虚质", "qixuzhi", "气虚血瘀：攻击回血+1。", "All", "Special", "Common", 0, "初始体质"); current_id += 1

    # --- 7. Special Cards: Enemies (As requested) ---
    # Act 1
    add_card(current_id, "风寒客", "fenghanshubiao", "寒邪侵袭：造成伤害并施加寒邪。", "EnemyDeck", "Special", "Common", 0, "敌人-Act1"); current_id += 1
    add_card(current_id, "风热袭", "ganhuochisheng", "热邪灼烧：多段伤害并施加热邪。", "EnemyDeck", "Special", "Common", 0, "敌人-Act1"); current_id += 1
    add_card(current_id, "湿浊缠", "pixushikun", "湿邪困脾：施加湿邪降低格挡。", "EnemyDeck", "Special", "Common", 0, "敌人-Act1"); current_id += 1
    add_card(current_id, "外感合病", "wuxingniluan", "精英：风寒/风热双态切换。", "EnemyDeck", "Special", "Elite", 0, "敌人-Act1-精英"); current_id += 1
    add_card(current_id, "风寒束表(Boss)", "fenghanshubiao", "Boss：寒凝血瘀，强力寒邪。", "EnemyDeck", "Special", "Boss", 0, "敌人-Act1-Boss"); current_id += 1
    add_card(current_id, "肝火炽盛(Boss)", "ganhuochisheng", "Boss：火旺伤阴，成长型攻击。", "EnemyDeck", "Special", "Boss", 0, "敌人-Act1-Boss"); current_id += 1
    
    # Act 2
    add_card(current_id, "气滞血瘀者", "qixuxueyu", "郁而作痛：施加气滞（增加消耗）和血瘀。", "EnemyDeck", "Special", "Common", 0, "敌人-Act2"); current_id += 1
    add_card(current_id, "脾虚湿盛者", "pixushikun", "湿困中焦：被动增加玩家卡牌消耗。", "EnemyDeck", "Special", "Common", 0, "敌人-Act2"); current_id += 1
    add_card(current_id, "心神不交者", "yinxuhuowang", "心悸不安：减少抽牌，概率晕眩。", "EnemyDeck", "Special", "Common", 0, "敌人-Act2"); current_id += 1
    add_card(current_id, "痰瘀互结", "pixushikun", "精英：根据玩家状态（湿/瘀）增强自身。", "EnemyDeck", "Special", "Elite", 0, "敌人-Act2-精英"); current_id += 1
    add_card(current_id, "脾虚湿困(Boss)", "pixushikun", "Boss：召唤水湿小怪，湿邪转化。", "EnemyDeck", "Special", "Boss", 0, "敌人-Act2-Boss"); current_id += 1

    # Act 3
    add_card(current_id, "阴阳离决者", "wuxingniluan", "阴阳格拒：双态切换，需特定时机攻击。", "EnemyDeck", "Special", "Common", 0, "敌人-Act3"); current_id += 1
    add_card(current_id, "冲任不固者", "qixuxueyu", "崩漏不止：清除玩家所有正面状态。", "EnemyDeck", "Special", "Common", 0, "敌人-Act3"); current_id += 1
    add_card(current_id, "厥阴复杂证", "wuxingniluan", "精英：寒热错杂，循环施加相反状态。", "EnemyDeck", "Special", "Elite", 0, "敌人-Act3-精英"); current_id += 1
    add_card(current_id, "五行失调(Boss)", "wuxingniluan", "最终Boss：五行流转，五阶段形态切换。", "EnemyDeck", "Special", "Boss", 0, "敌人-Act3-Boss"); current_id += 1

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CardIndex"
    
    # Write Headers
    ws.append(headers)
    
    # Style Headers
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Write Data
    for row in data:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save
    wb.save("CardMaster.xlsx")
    print("CardMaster.xlsx created successfully.")

if __name__ == "__main__":
    create_card_master_excel()
